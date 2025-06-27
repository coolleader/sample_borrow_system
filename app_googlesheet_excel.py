import streamlit as st
import pandas as pd
from datetime import datetime
import gspread
import io
from google.oauth2.service_account import Credentials
import openpyxl.utils.cell

# Google Sheets 授权
SCOPE = ["https://www.googleapis.com/auth/spreadsheets"]
creds = Credentials.from_service_account_info(
    st.secrets["gcp_service_account"],
    scopes=SCOPE
)
gc = gspread.authorize(creds)

# Google Sheet 配置
SHEET_ID = "14NUgJ8kk9DJtWaRKtIM_bQ5VXTJkZDibG28z_v_AwbU"
SHEET_NAME = "sample_inventory"
worksheet = gc.open_by_key(SHEET_ID).worksheet(SHEET_NAME)

# 表头字段
COLUMNS = [
    '型号', '序列号', '料号', '样品快递号', '状态',
    '送出时间', '送出客户', '送出附件',
    '收货时间', '收货快递号', '归还附件'
]

# 读取数据
def load_data():
    try:
        records = worksheet.get_all_records()
        if not records:
            return pd.DataFrame(columns=COLUMNS)
        return pd.DataFrame(records)
    except:
        return pd.DataFrame(columns=COLUMNS)

# 保存数据
def save_data(df):
    df = df.astype(str)
    worksheet.clear()
    worksheet.update([df.columns.tolist()] + df.values.tolist())

# UI 开始
df = load_data()
st.title("📦 样品送存管理系统")

menu = ["样品登记", "送出样品", "归还样品", "当前状态", "删除样品"]
choice = st.radio("选择操作", menu)

if choice == "样品登记":
    st.header("📄 样品登记")
    sample_type = st.text_input("型号").strip()
    sample_id = st.text_input("序列号").strip()
    sample_material = st.text_input("料号").strip()
    sample_deliver_id = st.text_input("样品快递号").strip()

    if st.button("登记"):
        if sample_id and sample_id not in df['序列号'].astype(str).values:
            new_row = pd.DataFrame([{
                '型号': sample_type,
                '序列号': sample_id,
                '料号': sample_material,
                '样品快递号': sample_deliver_id,
                '状态': '在库',
                '送出时间': '', '送出客户': '', '送出附件': '',
                '收货时间': '', '收货快递号': '', '归还附件': ''
            }])
            df = pd.concat([df, new_row], ignore_index=True)
            save_data(df)
            st.success("✅ 样品已登记")
        else:
            st.warning("⚠️ 序列号为空或已存在")

elif choice == "送出样品":
    st.header("📤 送出样品")
    sid = st.text_input("序列号").strip()
    client = st.text_input("送出客户").strip()
    send_attach = st.text_input("送出附件").strip()

    if st.button("确认送出"):
        if sid in df['序列号'].astype(str).values:
            idx = df[df['序列号'].astype(str) == sid].index[0]
            if df.at[idx, '状态'] == '在库':
                df.at[idx, '状态'] = '送出'
                df.at[idx, '送出时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                df.at[idx, '送出客户'] = client
                df.at[idx, '送出附件'] = send_attach
                df.at[idx, '收货时间'] = ''
                df.at[idx, '收货快递号'] = ''
                df.at[idx, '归还附件'] = ''
                save_data(df)
                st.success("✅ 样品送出成功")
            else:
                st.warning("⚠️ 样品不是在库状态")
        else:
            st.warning("⚠️ 样品不存在")

elif choice == "归还样品":
    st.header("📥 归还样品")
    sid = st.text_input("序列号").strip()
    deliver_id = st.text_input("收货快递号").strip()
    return_attach = st.text_input("归还附件").strip()

    if st.button("确认归还"):
        if sid in df['序列号'].astype(str).values:
            idx = df[df['序列号'].astype(str) == sid].index[0]
            if df.at[idx, '状态'] == '送出':
                df.at[idx, '状态'] = '在库'
                df.at[idx, '收货时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                df.at[idx, '收货快递号'] = deliver_id
                df.at[idx, '归还附件'] = return_attach
                save_data(df)
                st.success("✅ 样品已归还")
            else:
                st.warning("⚠️ 样品不是送出状态")
        else:
            st.warning("⚠️ 样品不存在")
            
elif choice == "当前状态":
    st.header("📊 当前样品状态")

    # 拷贝副本用于前端展示
    df_display = df.copy()

    # ✅ 将所有列都转为字符串并添加 '\t' 前缀，防止格式化
    df_display = df_display.applymap(lambda x: f"\t{x}" if pd.notnull(x) else "")

    # 显示表格，所有列均为文本格式
    st.dataframe(df_display, use_container_width=True)

    # ✅ 保留原始 df 导出 Excel，显式设置所有列为文本格式
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='样品数据')
        ws = writer.sheets['样品数据']
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
            for cell in ws[col_letter]:
                cell.number_format = '@'  # 设置为文本格式
    excel_buffer.seek(0)

    st.download_button(
        label="📥 下载样品表 (Excel)",
        data=excel_buffer,
        file_name="样品表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

elif choice == "删除样品":
    st.header("❌ 删除样品")
    sid = st.text_input("要删除的序列号").strip()
    confirm = st.checkbox("确认删除该样品")

    if st.button("删除"):
        if sid in df['序列号'].astype(str).values:
            if confirm:
                df = df[df['序列号'].astype(str) != sid]
                save_data(df)
                st.success("✅ 样品已删除")
            else:
                st.warning("⚠️ 请勾选确认删除")
        else:
            st.warning("⚠️ 样品不存在")
